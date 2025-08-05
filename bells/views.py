from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import BellSchedule
import uuid
from django.db.models import Min,Max
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

def bells_view(request):
    days = {
        'monday': 'Понедельник',
        'tuesday': 'Вторник',
        'wednesday': 'Среда',
        'thursday': 'Четверг',
        'friday': 'Пятница',
        'saturday': 'Суббота',
        'sunday': 'Воскресенье',
    }
    
    active_day = request.GET.get('day') or request.session.get('last_selected_day', 'monday')
    request.session['last_selected_day'] = active_day
    current_day = days.get(active_day, 'Понедельник')
    schedule = BellSchedule.objects.filter(day=active_day).order_by('order')
    
    # Получаем статистику по дням
    day_stats = {}
    for day in days:
        bells = BellSchedule.objects.filter(day=day)
        count = bells.count()
        if count > 0:
            first = bells.aggregate(Min('start_time'))['start_time__min']
            last = bells.aggregate(Max('end_time'))['end_time__max']
            day_stats[day] = {
                'count': count,
                'first': first.strftime('%H:%M'),
                'last': last.strftime('%H:%M')
            }
        else:
            day_stats[day] = {'count': 0, 'first': '', 'last': ''}
    
    return render(request, 'bells/index.html', {
        'days': days,
        'day_stats': day_stats,
        'active_day': active_day,
        'current_day_name': current_day,
        'schedule': schedule,
        'melodies': ['Звонок 1', 'Звонок 2', 'Звонок 3']
    })

@csrf_exempt
def save_bells(request):
    if request.method == 'POST':
        day = request.POST.get('day')
        request.session['last_selected_day'] = day
        # Удаляем старые записи для этого дня
        BellSchedule.objects.filter(day=day).delete()
        
        # Получаем данные из формы
        bells_data = request.POST.getlist('bells[]')
        
        for i, bell_data in enumerate(bells_data):
            data = bell_data.split('|')
            BellSchedule.objects.create(
                id=uuid.uuid4(),
                day=day,
                start_time=data[0],
                end_time=data[1],
                melody=data[2],
                is_active=data[3] == 'true',
                order=i
            )
        
        return JsonResponse({
            'status': 'success',
            'redirect_url': f'/bells/?day={day}'
            })
    return JsonResponse({'status': 'error'})

@csrf_exempt
def delete_bell(request, bell_id):
    try:
        bell = BellSchedule.objects.get(id=bell_id)
        bell.delete()
        return JsonResponse({'status': 'success'})
    except BellSchedule.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Звонок не найден'})
    
def get_schedule(request):
    day = request.GET.get('day', 'monday')
    schedule = BellSchedule.objects.filter(day=day).order_by('order')
    
    data = {
        'day_name': 'Понедельник',
        'schedule': [
            {
                'id': str(bell.id),
                'start_time': bell.start_time.strftime('%H:%M'),
                'end_time': bell.end_time.strftime('%H:%M'),
                'melody': bell.melody,
                'is_active': bell.is_active,
                'order': bell.order
            }
            for bell in schedule
        ]
    }
    
    return JsonResponse(data)

def export_to_excel(request):
    # Создаем новую книгу Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание звонков"

    # Настройки стилей
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=12)
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовки столбцов
    headers = ["День недели", "Начало", "Конец", "Мелодия", "Активен", "Порядок"]
    ws.append(headers)

    # Форматируем заголовки
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Порядок дней недели
    day_order = {
        'monday': 0, 'tuesday': 1, 'wednesday': 2,
        'thursday': 3, 'friday': 4, 'saturday': 5, 'sunday': 6
    }

    # Получаем и сортируем данные
    bells = BellSchedule.objects.all().order_by('order')
    bells_sorted = sorted(bells, key=lambda x: day_order[x.day])

    # Заполняем данные
    for bell in bells_sorted:
        row = [
            bell.get_day_display(),
            bell.start_time.strftime('%H:%M'),
            bell.end_time.strftime('%H:%M'),
            bell.melody,
            "Да" if bell.is_active else "Нет",
            bell.order
        ]
        ws.append(row)
        
        # Форматируем данные
        for cell in ws[ws.max_row]:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

    # Автоподгонка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Фиксируем заголовки
    ws.freeze_panes = 'A2'

    wb.save(output)
    output.seek(0)

    #  Создаем HttpResponse с правильными заголовками
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bells_schedule.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

@csrf_exempt
def clone_schedule(request):
    try:
        source_day = request.POST.get('source_day')
        target_days = request.POST.getlist('target_days')  # Исправлено получение target_days
        overwrite = request.POST.get('overwrite', 'false') == 'true'
        
        if not source_day or source_day not in dict(BellSchedule.DAYS_OF_WEEK).keys():
            raise ValueError("Неверный исходный день")
            
        source_schedule = BellSchedule.objects.filter(day=source_day).order_by('order')
        
        if not source_schedule.exists():
            return JsonResponse({
                'status': 'error', 
                'message': 'В исходном дне нет расписания для клонирования'
            }, status=400)
        
        created_count = 0
        for day in target_days:
            if day not in dict(BellSchedule.DAYS_OF_WEEK).keys():
                continue
                
            if overwrite:
                BellSchedule.objects.filter(day=day).delete()
            
            for bell in source_schedule:
                BellSchedule.objects.create(
                    id=uuid.uuid4(),
                    day=day,
                    start_time=bell.start_time,
                    end_time=bell.end_time,
                    melody=bell.melody,
                    is_active=bell.is_active,
                    order=bell.order
                )
                created_count += 1
        
        day_names = ', '.join([dict(BellSchedule.DAYS_OF_WEEK)[day] for day in target_days])
        return JsonResponse({
            'status': 'success',
            'message': f'Расписание успешно клонировано на: {day_names}'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=400)